# -*- coding: utf-8 -*-
__author__ = 'HoPun'

import os
from tkinter import *
from tkinter.filedialog import askdirectory
import tkinter
import openpyxl
import time

def file_name(file_dir):
    L=[]
    for root, dirs, files in os.walk(file_dir):
        print(files)
        for file in files:
            L.append(file)
    return L

def read_license(license_dir):
    list = []
    for root, dirs, files in os.walk(license_dir):
        # print(files)
        for file in files:
            with open(os.path.join(root, file),errors='ignore') as f:
                line = f.readline()
                # print(line)
                list.append(line)
                f.close()
    return list

def write_excel(license_list,output_path):


    license_4k_list = license_list[:4000]
    print(len(license_4k_list))

    # excel = xlwt.Workbook(encoding='utf-8', style_compression=0)
    # sheet = excel.add_sheet('Sheet', cell_overwrite_ok=True)
    # sheet.write(0, 0, 'deviceId')
    excel = openpyxl.Workbook()
    sheet = excel.active
    # sheet = excel.create_sheet()
    # sheet.title = 'Sheet'
    sheet.cell(1, 1, 'deviceId')
    i = 2
    for row in license_4k_list:
        sheet.cell(i, 1, row)
        i = i + 1
    cur_time_rec = time.localtime()
    nowTime = int(round(time.time() * 1000))

    excel.save(output_path + "/zibot{0:04d}{1:02d}{2:02d}-{3:02d}{4:02d}{5:02d}{6:20d}_speech_sbcAI.xlsx".format(
        cur_time_rec.tm_year, cur_time_rec.tm_mon, cur_time_rec.tm_mday,
        cur_time_rec.tm_hour, cur_time_rec.tm_min, cur_time_rec.tm_sec,nowTime))

    if(len(license_list)>4000):
        new_license_list = license_list[4000:]
        write_excel(new_license_list,output_path)


class selectPath():
    def __init__(self):
        self.path = StringVar()

    def set_path(self):
        self.path.set(askdirectory())
        # print(self.path.get())

    def get_path(self):
        if(self.path !=None):
            return self.path

    def get(self):
        print(self.path.get())
        return self.path.get()

if __name__ == '__main__':
    root = Tk()
    root.title = "License2Excel"
    input_path = selectPath()
    output_path = selectPath()
    # Thinker总共提供了三种布局组件的方法：pack(),grid()和place()
    #  grid()方法允许你用表格的形式来管理组件的位置
    #  row选项代表行，column选项代表列
    #  例如row=1，column=2表示第二行第三列(0表示第一行)


    # 如果表格大于组件，那么可以使用sticky选项来设置组件的位置
    #  同样你需要使用N，E，S,W以及他们的组合NE，SE，SW，NW来表示方位
    tkinter.Button(root, text="Input", width=10, command = input_path.set_path).grid(row=0, column=4, sticky=W, padx=10, pady=5)
    tkinter.Button(root, text="Output", width=10, command = output_path.set_path).grid(row=1, column=4, sticky=W, padx=10, pady=5)

    tkinter.Button(root, text="GEN", width=10, command = lambda:write_excel(read_license(input_path.get()),output_path.get())).grid(row=2, column=1, sticky=E, padx=10, pady=5)

    tkinter.Entry(root, textvariable=input_path.get_path()).grid(row=0, column=1)
    tkinter.Entry(root, textvariable=output_path.get_path()).grid(row=1, column=1)

    mainloop()








